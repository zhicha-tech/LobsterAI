#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导入 Excel 数据到飞书多维表格（使用字段名）
"""

import requests
import json
import time
from openpyxl import load_workbook

# 飞书配置
APP_ID = "cli_a925788df1f99cd6"
APP_SECRET = "R9FXqSIoXLmRSzScOV1kecENLx00atkK"
APP_TOKEN = "Cxu9bMv5qaqp7xsrIrecuUCeneg"
TABLE_ID = "tblJE4bs9PT10eUX"

# 获取 access_token
def get_access_token():
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    headers = {"Content-Type": "application/json"}
    data = {"app_id": APP_ID, "app_secret": APP_SECRET}

    response = requests.post(url, headers=headers, json=data)
    result = response.json()

    if result.get('code') == 0:
        return result['tenant_access_token']
    else:
        raise Exception(f"获取 token 失败: {result}")


# 批量创建记录（每次最多500条）
def batch_create_records(access_token, records):
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{APP_TOKEN}/tables/{TABLE_ID}/records/batch_create"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    data = {"records": records}

    response = requests.post(url, headers=headers, json=data)
    return response.json()


# 读取 Excel 文件
file_path = r'c:\Users\Admin\Documents\xwechat_files\wxid_34h0ipa8hufm22_c3d1\msg\file\2025-09\2026年度内蒙古自治区事业单位公开招聘第一阶段未达到规定开考比例取消岗位（挂网）.xlsx'

print('正在读取 Excel 文件...')
wb = load_workbook(file_path, data_only=True)
ws = wb.active

# 读取列名（第2行）
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(2, col).value
    headers.append(header if header else f'列{col}')

print(f'列名: {headers}')

# 读取所有数据（从第3行开始）
all_data = []
for row in range(3, ws.max_row + 1):
    row_dict = {}
    for col in range(1, len(headers) + 1):
        header = headers[col - 1]
        cell_value = ws.cell(row, col).value
        row_dict[header] = cell_value
    all_data.append(row_dict)

print(f'共读取 {len(all_data)} 条记录')
wb.close()

# 获取 access_token
print('\n正在连接飞书多维表格...')
access_token = get_access_token()
print('✅ 获取 access_token 成功')

# 准备导入数据
print(f'\n开始导入数据到多维表格...')
print(f'多维表格链接: https://vantasma.feishu.cn/base/{APP_TOKEN}')

# 构建记录列表
records_to_import = []
for i, record in enumerate(all_data):
    # 使用字段名构建数据
    fields = {}

    # 职位代码
    职位代码 = str(record.get('职位代码', ''))
    if 职位代码:
        fields["职位代码"] = 职位代码

    # 单位名称
    单位名称 = record.get('单位名称', '')
    if 单位名称:
        fields["单位名称"] = 单位名称

    # 职位名称
    职位名称 = record.get('职位名称', '')
    if 职位名称:
        fields["职位名称"] = 职位名称

    # 招考人数
    招考人数 = record.get('招考人数', 0)
    if 招考人数:
        fields["招考人数"] = 招考人数

    # 交费人数
    交费人数 = record.get('交费人数', 0)
    if 交费人数:
        fields["交费人数"] = 交费人数

    # 备注
    备注 = record.get('备注', '取消')
    if 备注:
        fields["备注"] = 备注

    # 数据来源
    fields["数据来源"] = "2026年度内蒙古自治区事业单位公开招聘第一阶段未达到规定开考比例取消岗位（挂网）"

    # 导入时间
    fields["导入时间"] = int(time.time() * 1000)

    records_to_import.append({"fields": fields})

# 批量导入（每批500条）
batch_size = 500
total_batches = (len(records_to_import) + batch_size - 1) // batch_size

success_count = 0
error_count = 0

for batch_num in range(total_batches):
    start_idx = batch_num * batch_size
    end_idx = min((batch_num + 1) * batch_size, len(records_to_import))
    batch_records = records_to_import[start_idx:end_idx]

    print(f'\n批次 {batch_num + 1}/{total_batches}: 导入 {len(batch_records)} 条记录...')

    try:
        result = batch_create_records(access_token, batch_records)

        if result.get('code') == 0:
            records_result = result.get('data', {}).get('records', [])
            success_count += len(records_result)
            print(f'  ✅ 成功导入 {len(records_result)} 条记录')
        else:
            error_count += len(batch_records)
            print(f'  ❌ 批次 {batch_num + 1} 失败: {result.get("msg", "Unknown")}')
    except Exception as e:
        error_count += len(batch_records)
        print(f'  ❌ 批次 {batch_num + 1} 错误: {e}')

    # 添加延迟，避免限流
    if batch_num < total_batches - 1:
        time.sleep(0.5)

print('\n' + '=' * 80)
print('导入完成！')
print('=' * 80)
print(f'✅ 成功导入: {success_count} 条')
print(f'❌ 失败: {error_count} 条')
print(f'\n📎 多维表格链接: https://vantasma.feishu.cn/base/{APP_TOKEN}')
