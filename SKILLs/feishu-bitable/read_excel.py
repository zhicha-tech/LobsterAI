#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""正确解析 Excel 文件"""

from openpyxl import load_workbook
import json

# 读取 Excel 文件
file_path = r'c:\Users\Admin\Documents\xwechat_files\wxid_34h0ipa8hufm22_c3d1\msg\file\2025-09\2026年度内蒙古自治区事业单位公开招聘第一阶段未达到规定开考比例取消岗位（挂网）.xlsx'

print('正在读取 Excel 文件...')
wb = load_workbook(file_path, data_only=True)
ws = wb.active

print(f'\n总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')

# 第1行是标题，跳过
# 第2行是列名
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(2, col).value
    if header:
        headers.append(header)
    else:
        headers.append(f'列{col}')

print(f'\n列名: {headers}')

# 从第3行开始读取数据
all_data = []
for row in range(3, ws.max_row + 1):
    row_dict = {}
    for col in range(1, len(headers) + 1):
        header = headers[col - 1]
        cell_value = ws.cell(row, col).value
        row_dict[header] = cell_value
    all_data.append(row_dict)

print(f'\n共读取 {len(all_data)} 条记录')

# 打印前10条记录
print('\n' + '=' * 80)
print('前10条记录预览:')
print('=' * 80)
for i, record in enumerate(all_data[:10], 1):
    print(f'\n记录 {i}:')
    for key, value in record.items():
        print(f'  {key}: {value}')

wb.close()

# 保存数据到 JSON 文件
output_file = 'excel_data.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump({
        'headers': headers,
        'total_rows': len(all_data),
        'data': all_data
    }, f, ensure_ascii=False, indent=2)

print(f'\n\n✅ 数据已保存到: {output_file}')

# 统计信息
print('\n' + '=' * 80)
print('统计信息:')
print('=' * 80)
print(f'总记录数: {len(all_data)}')
print(f'列数: {len(headers)}')
print(f'列名: {", ".join(headers)}')

# 统计取消原因
cancel_reasons = {}
for record in all_data:
    reason = record.get('备注', '未知')
    cancel_reasons[reason] = cancel_reasons.get(reason, 0) + 1

print('\n取消原因统计:')
for reason, count in sorted(cancel_reasons.items(), key=lambda x: x[1], reverse=True):
    print(f'  {reason}: {count} 条')
